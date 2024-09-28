from datetime import datetime, timedelta
import os
from tkinter import filedialog, messagebox
from typing import Literal

import gspread
from gspread.exceptions import SpreadsheetNotFound
from gspread.spreadsheet import Spreadsheet
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet as ExcelWorksheet
import requests

from selenium_youtube.constants import UPLOAD_POLL_FREQUENCY
import selenium_youtube.upload as selytupload
from utils.datetime import compare_datetimes_lazily
from utils.path import list_all_files, get_project_directory
from utils.settings import get_setting, InvalidSettingsError
from utils.threads import wait_until_number_of_threads_is
import valorant_autotracker.constants as c

def get_google_sheet(name: str) -> Spreadsheet:
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(get_setting(*c.GOOGLE_SERVICE_ACCOUNT_KEY_JSON_PATH_LOCATOR), c.SCOPE)
    except FileNotFoundError as e:
        raise FileNotFoundError("Serivce account key not found. Please make sure you have set a valid service account key path (.json) in the settings.") from e
    except ValueError as e:
        raise c.InvalidGoogleServiceAccountKeyError("Credential type is not SERVICE_ACCOUNT. Please make sure your key is for a service account.") from e
    except KeyError as e:
        raise c.InvalidGoogleServiceAccountKeyError("Expected key(s) not present. Please make sure your service account key is valid.") from e

    client = gspread.authorize(creds)

    try:
        sheet = client.open(name)
    except SpreadsheetNotFound as e:
        raise SpreadsheetNotFound("Spreadsheet not found. Please make sure you have set a valid spreadsheet name in the settings.") from e

    return sheet

def get_excel_workbook(path: str) -> Workbook:
    try:
        workbook = load_workbook(path)
    except InvalidFileException as e:
        raise InvalidFileException("Invalid file extension. Please make sure you your excel file has one of the supported extensions (.xlsx, .xlsm, .xltx, or .xltm) in the settings.") from e
    except FileNotFoundError as e:
        raise FileNotFoundError("Excel file not found. Please make sure you have set a valid excel file path in the settings.") from e

    return workbook

def insert_row_to_excel_sheet(workbook: Workbook, path: str, row: int, values: list) -> None:
    sheet: ExcelWorksheet = workbook.active

    if sheet is None:
        raise FileNotFoundError("Excel sheet not found. Please make sure your excel workbook contains atleast 1 sheet.")

    sheet.insert_rows(row) # Inserts empty row

    sheet._current_row = row - 1 # Sets writer's final row to row above, so appended values ago to row
    sheet.append(values)

    workbook.save(path)

def append_row_to_excel_sheet(workbook: Workbook, path: str, values: list) -> None:
    sheet: ExcelWorksheet = workbook.active

    if sheet is None:
        raise FileNotFoundError("Excel sheet not found. Please make sure your excel workbook contains atleast 1 sheet.")

    sheet.append(values)

    workbook.save(path)

def fill_cells(workbook: Workbook, path: str, cell_range: str, pattern_fill: PatternFill) -> None:
    sheet: ExcelWorksheet = workbook.active

    if sheet is None:
        raise FileNotFoundError("Excel sheet not found. Please make sure your excel workbook contains atleast 1 sheet.")

    for cell in sheet[cell_range]:
        cell.fill = pattern_fill

    workbook.save(path)

def make_default_excel_file(path: str) -> None:
    workbook = Workbook()

    spreadsheet_format_setting = get_setting(*c.SPREADSHEET_FORMAT_LOCATOR).split(",")
    column_headers = [c.SPREADSHEET_FORMAT_OPTIONS[column_header] for column_header in spreadsheet_format_setting] # Converts into user-friendly format
    insert_row_to_excel_sheet(workbook, path, 1, column_headers)

    # Makes row yellow
    pattern_fill = PatternFill(start_color="FFFF00", fill_type="solid")
    fill_cells(workbook, path, "1:1", pattern_fill)

class ValorantAPI:
    def __init__(self, api_key) -> None:
        self.headers = {"Authorization": api_key}

    def manage_henrikdev_api_errors(self, status_code: int) -> Literal[True]:
        match status_code:
            case 200:
                return True
            case 400:
                raise c.APIError("Request error by the client. Please make sure you have set a valid PUUID and region in the settings.")
            case 401 | 403:
                raise c.APIError("Invalid API Key. Please make sure you have a valid API key in the settings.")
            case 404:
                raise c.APIError("Player not found, invalid PUUID. Please make sure you have set a valid PUUID and region in the settings.")
            case 408:
                raise c.APIError("Timeout while fetching riot data. Please try again.")
            case 410:
                raise c.APIError("Endpoint is deprecated. Please contact the developer.")
            case 429:
                raise c.APIError("Rate limit reached. Please try again later.")
            case 500:
                raise c.APIError("Internal server error. Please make sure you have set a valid PUUID and region in the settings.")
            case 501:
                raise c.APIError("API Version not implemented. Please contact the developer.")
            case 503:
                raise c.APIError("Riot API seems to be down, API unable to connect. Please try again later.")
            case _:
                raise c.APIError(f"An error ({status_code}) has occured. Please try again or report this to 'github.com/aritra-codes/valorant-autotracker'.")

    def send_request(self, url: str, params: dict | list[tuple] | None=None) -> dict:
        response = requests.get(url, params, headers=self.headers, timeout=c.API_REQUEST_TIMEOUT)

        self.manage_henrikdev_api_errors(response.status_code)

        return response.json()

    def find_puuid(self, name: str, tag: str) -> str:
        response = self.send_request(c.ACCOUNT_BY_NAME_URL(name, tag))

        return response["data"]["puuid"]

    def find_region(self, name: str, tag: str) -> str:
        response = self.send_request(c.ACCOUNT_BY_NAME_URL(name, tag))

        return response["data"]["region"]

    def get_matches(self, puuid: str, affinity: c.Affinity, size: int=10) -> list:
        response = self.send_request(c.MATCHES_URL(puuid, affinity), {"mode": "competitive", "size": size})

        matches = response["data"]
        matches.reverse() # Reverses from descending to ascending

        return matches

    def get_new_matches(self, puuid: str, affinity: c.Affinity, latest_match_id: str) -> list:
        matches = self.get_matches(puuid, affinity)

        for index, match in enumerate(reversed(matches)):
            if match["metadata"]["matchid"] == latest_match_id:
                return matches[::-1][:index][::-1] # Reverses to descending, gets index number of matches, reverses back to ascending

        upload_last_10 = messagebox.askokcancel("Match not found",
                                                "No match in the last ~10 matches has the latest match ID set in the settings."
                                                " Would you like to insert and/or upload the last ~10 matches?")

        if upload_last_10:
            return matches

        raise c.MatchNotFoundError("No match in the last ~10 matches has the latest match ID set in the settings.")

    def get_mmr_changes(self, puuid: str, affinity: c.Affinity, size: int=10) -> dict[str, int]:
        response = self.send_request(c.MMR_HISTORY_URL(puuid, affinity), {"size": size})

        mmr_changes = {match["match_id"]: match["mmr_change_to_last_game"] for match in response["data"]}

        return mmr_changes

def convert_valorant_date(unformatted_date: str) -> datetime:
    datetime_obj = datetime.strptime(unformatted_date, c.VALORANT_DATE_FORMAT)
    return datetime_obj + timedelta(hours=1, minutes=2) # Accounts for delay time

def format_match_info(match_info: dict, puuid: str, mmr_change: int) -> dict[str, str | int]:
    meta = match_info["metadata"]
    player = next(player for player in match_info["players"]["all_players"] if player["puuid"] == puuid)
    stats = player["stats"]
    team = match_info["teams"][player["team"].lower()]

    date_started_datetime = convert_valorant_date(meta["game_start_patched"])
    headshot_percentage = round((stats["headshots"] / (stats["headshots"] + stats["bodyshots"] + stats["legshots"])) * 100)
    average_damage_per_round = round(player["damage_made"] / meta["rounds_played"])

    use_mdy_dates = get_setting(*c.USE_MDY_DATES_SETTING_LOCATOR, boolean=True)

    formatted_match_info = {"match_id": meta["matchid"],
                            "date_started": date_started_datetime.strftime(r"%m/%d/%Y" if use_mdy_dates else r"%d/%m/%Y"),
                            "rank": player["currenttier_patched"],
                            "mmr_change": mmr_change,
                            "rounds_won": team["rounds_won"], 
                            "rounds_lost": team["rounds_lost"], 
                            "tracker_link": f"https://tracker.gg/valorant/match/{meta["matchid"]}",
                            "map": meta["map"],
                            "agent": player["character"],
                            "kills": stats["kills"],
                            "deaths": stats["deaths"],
                            "assists": stats["assists"],
                            "headshot_percentage": headshot_percentage,
                            "average_damage_per_round": average_damage_per_round}

    link = ""

    if get_setting(*c.AUTOUPLOAD_VIDEOS_SETTING_LOCATOR, boolean=True):
        default_number_of_threads = get_setting(*c.DEFAULT_NUMBER_OF_THREADS, integer=True)
        max_videos_simultaneously = get_setting(*c.MAX_VIDEOS_SIMULTANEOUSLY_SETTING_LOCATOR, integer=True)

        # Waits for number of videos uploading to reach setting
        wait_until_number_of_threads_is(default_number_of_threads + (max_videos_simultaneously - 1), UPLOAD_POLL_FREQUENCY)

        try:
            link = upload_video(formatted_match_info, date_started_datetime)
        except (FileNotFoundError, c.VideoUploadError) as e:
            print(f"{e} Skipping upload...")

    formatted_match_info["video_link"] = link

    return formatted_match_info

def upload_video(match_info: dict[str, str | int], date_started_datetime: datetime) -> str:
    firefox_profile_path = get_setting(*c.FIREFOX_PROFILE_SETTING_LOCATOR)
    title = format_video_title(match_info)
    background = get_setting(*c.BACKGROUND_PROCESS_SETTING_LOCATOR, boolean=True)

    try:
        visibility = c.Visibility[get_setting(*c.VIDEO_VISIBILITY_SETTING_LOCATOR)]
    except KeyError as e:
        raise InvalidSettingsError("'visibility' setting is not valid. Please check and save your settings.") from e

    if get_setting(*c.AUTOSELECT_VIDEOS_SETTING_LOCATOR, boolean=True):
        try:
            video_path = autofind_video_path(date_started_datetime)
        except FileNotFoundError:
            video_path = input_file_path(f"Autofind could not find the video, open video for '{title}'")
    else:
        video_path = input_file_path(f"Open video for '{title}'")


    if video_path:
        print(f"Starting upload of video '{title}'...")

        attempts = 2

        while True:
            try:
                video_link = selytupload.upload_video(firefox_profile_path,
                                                video_path,
                                                title,
                                                visibility=visibility,
                                                background=background)
            except FileNotFoundError as e:
                raise c.FirefoxProfileNotFoundError("Firefox profile not found. Please make sure you have set a valid firefox profile path in the settings.") from e
            except Exception as e:
                """
                WIP: Allow user to decide whether to reupload video.

                reupload = input_messagebox_yes_or_no("Reupload Video?", f"An error has occured, would you like to reupload the video for '{title}'?")

                if not reupload:
                    print(f"Skipping upload for '{title}'...")

                    break
                """

                attempts -= 1

                if attempts <= 0:
                    raise c.VideoUploadError(f"Video '{title}' failed to upload.") from e

                print(f"Video '{title}' failed to upload. Retrying...")
            else:
                return video_link

    raise FileNotFoundError(f"No video path found for '{title}'.")

def input_messagebox_yes_or_no(title: str, message: str) -> bool:
    msgbox = messagebox.askquestion(title, message)

    return msgbox == "yes"

def input_file_path(title: str) -> str:
    file_path = filedialog.askopenfilename(title=title)

    return file_path

def autofind_video_path(match_start_datetime: datetime) -> str:
    recording_start_delay = get_setting(*c.RECORDING_START_DELAY_SETTING_LOCATOR, floatp=True)
    match_start_datetime += timedelta(seconds=recording_start_delay)

    directory = get_setting(*c.VIDEO_DIRECTORY_SETTING_LOCATOR)
    filename_format = get_setting(*c.FILENAME_FORMAT_SETTING_LOCATOR)

    try:
        for filename in reversed(list_all_files(directory)):
            # Checks if filename follows format setting
            try:
                video_datetime = datetime.strptime(filename, filename_format)
            except ValueError:
                continue

            # Checks if datetime from filename matches
            if (compare_datetimes_lazily(video_datetime, match_start_datetime) or
                compare_datetimes_lazily(video_datetime, (match_start_datetime - timedelta(minutes=1))) or
                compare_datetimes_lazily(video_datetime, (match_start_datetime + timedelta(minutes=1)))):
                return os.path.join(directory, filename)
    except FileNotFoundError as e:
        raise c.VideoDirectoryNotFoundError("Video directory not found. Please make sure you have set a valid video directory in the settings.") from e

    raise FileNotFoundError

def format_video_title(formatted_match_info: dict) -> str:
    return f"VALORANT {formatted_match_info["date_started"]} {formatted_match_info["map"]} {formatted_match_info["agent"]} {formatted_match_info["rank"].replace(" ", "")}"
